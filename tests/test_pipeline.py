import csv
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from income_book_automation.exporters.income_book import HelperColumnMapping
from income_book_automation.models import (
    BankName,
    BankTransaction,
    CheckboxPaymentMethod,
    ClientProfile,
    ReviewField,
    TransactionCategory,
    TransactionSource,
)
from income_book_automation.parsers.checkbox import (
    MissingCheckboxColumnError,
)
from income_book_automation.parsers.errors import BankStatementFormatError
from income_book_automation.pipeline import (
    BankStatementSource,
    IncomeBookPipelineError,
    MissingStatementAccountError,
    MixedPeriodError,
    SenseAcquiringRequiresCheckboxError,
    UnresolvedTransactionsError,
    UnsupportedCurrencyError,
    run_income_book_pipeline,
)


def _write_pumb_statement(
    path: Path,
    *,
    include_incomplete_transaction: bool = False,
    first_payment_purpose: str = "Оплата за тестові послуги",
) -> None:
    rows = [
        [
            "ST_DATE",
            "ACC_NUMB",
            "CUR_NUMB",
            "DOC_NO",
            "DB",
            "CR",
            "KOR_NAME",
            "KOR_ACC",
            "KOR_OKPO",
            "DESCRIPT",
        ],
        [
            "2026.06.01",
            "UA273000010000000000000000001",
            "980",
            "TEST-001",
            "0",
            "20.00",
            "ТОВ Тестовий покупець",
            "UA753000010000000000000000010",
            "11111111",
            first_payment_purpose,
        ],
        [
            "2026.06.01",
            "UA273000010000000000000000001",
            "980",
            "TEST-002",
            "0",
            "30.00",
            "ФОП Тестовий Тарас Іванович",
            "UA973000010000000000000000002",
            "0000000000",
            "Переказ між власними рахунками",
        ],
    ]

    if include_incomplete_transaction:
        rows.append(
            [
                "2026.06.01",
                "UA273000010000000000000000001",
                "980",
                "TEST-003",
                "0",
                "40.00",
                "",
                "",
                "",
                "",
            ]
        )

    with path.open("w", encoding="cp1251", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerows(rows)


def _write_overlapping_pumb_statement(
    path: Path,
    *,
    unique_document_number: str,
    unique_amount: str,
) -> None:
    rows = [
        [
            "ST_DATE",
            "ACC_NUMB",
            "CUR_NUMB",
            "DOC_NO",
            "DB",
            "CR",
            "KOR_NAME",
            "KOR_ACC",
            "KOR_OKPO",
            "DESCRIPT",
        ],
        [
            "2026.06.01",
            "UA273000010000000000000000001",
            "980",
            "TEST-DUPLICATE-001",
            "0",
            "20.00",
            "ТОВ Тестовий покупець",
            "UA753000010000000000000000010",
            "11111111",
            "Оплата за тестові послуги",
        ],
        [
            "2026.06.01",
            "UA273000010000000000000000001",
            "980",
            unique_document_number,
            "0",
            unique_amount,
            "ТОВ Інший тестовий покупець",
            "UA483000010000000000000000011",
            "22222222",
            "Інша тестова оплата",
        ],
    ]

    with path.open("w", encoding="cp1251", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerows(rows)


def _write_abank_statement(path: Path) -> None:
    rows = [
        [
            "Дата операції",
            "Час операції",
            "№ платежу",
            "Тип операції",
            "Контрагент",
            "ЄДРПОУ/РНОКПП контрагента",
            "IBAN Контрагента",
            "Призначення платежу",
            "Сума, грн",
            "Залишок після операції в валюті рахунку",
        ],
        [
            "01.06.2026",
            "12:00:00",
            "TEST-ABANK-001",
            "Вхідна",
            "ТОВ Тестовий покупець",
            "11111111",
            "UA753000010000000000000000010",
            "Оплата за тестові послуги",
            "20,00",
            "20,00",
        ],
    ]

    with path.open("w", encoding="utf-8-sig", newline="") as file:
        file.write(
            "Виписка за рахунком ФОП ТЕСТОВИЙ ТАРАС ІВАНОВИЧ "
            "UA273000010000000000000000001 UAH "
            "за період з 01.06.2026-30.06.2026\n"
        )
        writer = csv.writer(file, delimiter=",")
        writer.writerows(rows)


def _write_sense_statement(
    path: Path,
    *,
    counterparty: str = "ТОВ Тестовий покупець",
    counterparty_tax_id: str = "11111111",
    payment_purpose: str = "Оплата за послуги; без ПДВ",
    amount: str = "20,00",
) -> None:
    headers = [
        "Наш рахунок",
        "Наш IBAN",
        "Операція",
        "Рахунок",
        "IBAN",
        "МФО банку контрагента",
        "Найменування контрагента",
        "Код контрагента",
        "Призначення платежу",
        "Дата проведення",
        "Номер документа",
        "Сума",
        "Валюта",
        "Час проведення",
        "Дата документа",
        "Дата архівування",
        "Ід.код",
        "Найменування",
        "МФО",
    ]
    row = [
        "26000000000001",
        "UA273000010000000000000000001",
        "Кредит",
        "26000000000002",
        "UA753000010000000000000000010",
        "300001",
        counterparty,
        counterparty_tax_id,
        payment_purpose,
        "01.06.2026",
        "TEST-SENSE-001",
        amount,
        "UAH",
        "12:00:00",
        "01.06.2026",
        "01.06.2026",
        "0000000000",
        "ФОП Тестовий Тарас Іванович",
        "300001",
    ]
    path.write_text(
        f"{';'.join(headers)}\n{';'.join(row)}\n",
        encoding="cp1251",
    )


def _write_checkbox_report(
    path: Path,
    *,
    card_revenue: int = 100,
    card_refund: int = 10,
    cash_revenue: int = 50,
    cash_refund: int = 0,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Дата відкриття",
            "Виручка безготівка",
            "Повернення безготівка",
            "Виручка готівка",
            "Повернення готівка",
        ]
    )
    sheet.append(
        [
            # Excel stores timestamps without timezone information.
            datetime(2026, 6, 1, 12, 0),  # noqa: DTZ001
            card_revenue,
            card_refund,
            cash_revenue,
            cash_refund,
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_income_book_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026"
    sheet["A6"] = date(2026, 6, 1)
    sheet["A6"].number_format = "yyyy-mm-dd"
    workbook.save(path)
    workbook.close()


def _client_profile() -> ClientProfile:
    return ClientProfile(
        client_id="client-001",
        legal_name="ФОП Тестовий Тарас Іванович",
        tax_id="0000000000",
        own_accounts={
            "UA273000010000000000000000001",
            "UA973000010000000000000000002",
        },
    )


def test_run_income_book_pipeline_processes_sources_and_exports_workbook(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "statement.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_pumb_statement(statement_path)
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank=BankName.PUMB,
        bank_statement_path=statement_path,
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
        helper_columns=HelperColumnMapping(
            total=10,
            checkbox_card=12,
            checkbox_cash=13,
            bank_income=14,
        ),
    )

    assert result.output_path == output_path
    assert len(result.daily_entries) == 1
    assert result.daily_entries[0].date == date(2026, 6, 1)
    assert result.daily_entries[0].checkbox_card_income == Decimal("90.00")
    assert result.daily_entries[0].checkbox_cash_income == Decimal("50.00")
    assert result.daily_entries[0].bank_income == Decimal("20.00")

    assert [record.category for record in result.classified_transactions] == [
        TransactionCategory.INCOME,
        TransactionCategory.OWN_TRANSFER,
    ]
    assert result.needs_review == ()

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        assert sheet["B6"].value == 160
        assert sheet["J6"].value == "=L6+M6+N6"
        assert sheet["K6"].value is None
        assert sheet["L6"].value == 90
        assert sheet["M6"].value == 50
        assert sheet["N6"].value == 20
    finally:
        workbook.close()


def test_pipeline_supports_checkbox_without_bank_statements(tmp_path: Path) -> None:
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[],
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert result.classified_transactions == ()
    assert result.daily_entries[0].checkbox_card_income == Decimal("90.00")
    assert result.daily_entries[0].checkbox_cash_income == Decimal("50.00")
    assert result.daily_entries[0].bank_income == Decimal("0.00")


def test_pipeline_supports_bank_statements_without_checkbox(tmp_path: Path) -> None:
    statement_path = tmp_path / "statement.csv"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _write_pumb_statement(statement_path)
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[
            BankStatementSource(bank=BankName.PUMB, path=statement_path),
        ],
        checkbox_path=None,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert result.checkbox_warnings == ()
    assert result.daily_entries[0].checkbox_card_income == Decimal("0.00")
    assert result.daily_entries[0].checkbox_cash_income == Decimal("0.00")
    assert result.daily_entries[0].bank_income == Decimal("20.00")


def test_pipeline_rejects_request_without_any_income_source(tmp_path: Path) -> None:
    with pytest.raises(IncomeBookPipelineError, match="at least one income source"):
        run_income_book_pipeline(
            client=_client_profile(),
            bank_statements=[],
            checkbox_path=None,
            template_path=tmp_path / "template.xlsx",
            output_path=tmp_path / "output.xlsx",
            sheet_name="2026",
        )


def test_pipeline_blocks_export_when_transaction_needs_review(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "statement-with-missing-fields.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_pumb_statement(
        statement_path,
        include_incomplete_transaction=True,
    )
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    with pytest.raises(UnresolvedTransactionsError) as error_info:
        run_income_book_pipeline(
            client=_client_profile(),
            bank=BankName.PUMB,
            bank_statement_path=statement_path,
            checkbox_path=checkbox_path,
            template_path=template_path,
            output_path=output_path,
            sheet_name="2026",
        )

    assert not output_path.exists()

    review_records = error_info.value.records
    assert len(review_records) == 1

    review_record = review_records[0]
    assert review_record.category is TransactionCategory.NEEDS_REVIEW
    assert review_record.transaction.source.original_filename == (
        "statement-with-missing-fields.csv"
    )
    assert review_record.transaction.source.row_number == 4
    assert review_record.missing_fields == frozenset(
        {
            ReviewField.COUNTERPARTY,
            ReviewField.COUNTERPARTY_ACCOUNT,
            ReviewField.COUNTERPARTY_TAX_ID,
            ReviewField.PAYMENT_PURPOSE,
        }
    )


def test_pipeline_exports_negative_checkbox_results_with_warnings(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "statement.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_pumb_statement(statement_path)
    _write_checkbox_report(
        checkbox_path,
        card_revenue=10,
        card_refund=25,
        cash_revenue=5,
        cash_refund=10,
    )
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank=BankName.PUMB,
        bank_statement_path=statement_path,
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert output_path.exists()
    assert result.daily_entries[0].checkbox_card_income == Decimal("-15.00")
    assert result.daily_entries[0].checkbox_cash_income == Decimal("-5.00")
    assert len(result.checkbox_warnings) == 2

    card_warning, cash_warning = result.checkbox_warnings
    assert card_warning.payment_method is CheckboxPaymentMethod.CARD
    assert card_warning.result == Decimal("-15.00")
    assert cash_warning.payment_method is CheckboxPaymentMethod.CASH
    assert cash_warning.result == Decimal("-5.00")

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        assert sheet["B6"].value == 0
        assert sheet["J6"].value == "=K6+L6+M6"
        assert sheet["K6"].value == -15
        assert sheet["L6"].value == -5
        assert sheet["M6"].value == 20
    finally:
        workbook.close()


def test_pipeline_names_original_bank_file_and_selected_bank(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "uploaded-privat.csv"
    statement_path.write_text("not a PUMB statement", encoding="utf-8")

    with pytest.raises(
        BankStatementFormatError,
        match="uploaded-privat.csv.*ПУМБ",
    ):
        run_income_book_pipeline(
            client=_client_profile(),
            bank=BankName.PUMB,
            bank_statement_path=statement_path,
            checkbox_path=tmp_path / "ZReport.xlsx",
            template_path=tmp_path / "income-book.xlsx",
            output_path=tmp_path / "output.xlsx",
            sheet_name="2026",
        )


def test_pipeline_identifies_wrong_checkbox_z_report(tmp_path: Path) -> None:
    statement_path = tmp_path / "bank.csv"
    checkbox_path = tmp_path / "wrong-checkbox-report.xlsx"

    _write_pumb_statement(statement_path)

    workbook = Workbook()
    workbook.active.append(["Це інший звіт Checkbox"])
    workbook.save(checkbox_path)
    workbook.close()

    with pytest.raises(MissingCheckboxColumnError) as error_info:
        run_income_book_pipeline(
            client=_client_profile(),
            bank=BankName.PUMB,
            bank_statement_path=statement_path,
            checkbox_path=checkbox_path,
            template_path=tmp_path / "income-book.xlsx",
            output_path=tmp_path / "output.xlsx",
            sheet_name="2026",
        )

    assert error_info.value.filename == "wrong-checkbox-report.xlsx"
    assert error_info.value.missing_headers == (
        "Дата відкриття",
        "Виручка безготівка",
        "Повернення безготівка",
        "Виручка готівка",
        "Повернення готівка",
    )
    assert "Книгу доходів не сформовано" in str(error_info.value)


def test_run_income_book_pipeline_requires_statement_account_for_mono(
    tmp_path: Path,
) -> None:
    with pytest.raises(
        MissingStatementAccountError,
        match="Mono «statement.csv» потрібно вказати IBAN",
    ):
        run_income_book_pipeline(
            client=_client_profile(),
            bank=BankName.MONO,
            bank_statement_path=tmp_path / "statement.csv",
            checkbox_path=tmp_path / "checkbox.xlsx",
            template_path=tmp_path / "template.xlsx",
            output_path=tmp_path / "output.xlsx",
            sheet_name="2026",
        )


def test_run_income_book_pipeline_supports_abank(tmp_path: Path) -> None:
    statement_path = tmp_path / "statement.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_abank_statement(statement_path)
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank=BankName.ABANK,
        bank_statement_path=statement_path,
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert len(result.classified_transactions) == 1
    assert result.classified_transactions[0].transaction.bank is BankName.ABANK
    assert result.daily_entries[0].bank_income == Decimal("20.00")


def test_run_income_book_pipeline_supports_sense_bank(tmp_path: Path) -> None:
    statement_path = tmp_path / "sense.csv"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_sense_statement(statement_path)
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[
            BankStatementSource(bank=BankName.SENSE, path=statement_path),
        ],
        checkbox_path=None,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert result.classified_transactions[0].transaction.bank is BankName.SENSE
    assert result.daily_entries[0].bank_income == Decimal("20.00")


def test_pipeline_excludes_sense_acquiring_when_checkbox_is_provided(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "sense-acquiring.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_sense_statement(
        statement_path,
        counterparty='АТ "СЕНС БАНК"',
        counterparty_tax_id="23494714",
        payment_purpose=(
            "Зарах.еквайрінг; сума 100.00грн; комісія 1.30грн"
        ),
        amount="98,70",
    )
    _write_checkbox_report(
        checkbox_path,
        card_revenue=100,
        card_refund=0,
        cash_revenue=0,
        cash_refund=0,
    )
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[
            BankStatementSource(bank=BankName.SENSE, path=statement_path),
        ],
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert (
        result.classified_transactions[0].category
        is TransactionCategory.EXCLUDED
    )
    assert result.daily_entries[0].checkbox_card_income == Decimal("100.00")
    assert result.daily_entries[0].bank_income == Decimal("0.00")
    assert result.daily_entries[0].total_income == Decimal("100.00")


def test_pipeline_blocks_sense_acquiring_without_checkbox(tmp_path: Path) -> None:
    statement_path = tmp_path / "sense-acquiring.csv"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_sense_statement(
        statement_path,
        counterparty='АТ "СЕНС БАНК"',
        counterparty_tax_id="23494714",
        payment_purpose=(
            "Зарах.еквайрінг; сума 100.00грн; комісія 1.30грн"
        ),
        amount="98,70",
    )
    _write_income_book_template(template_path)

    with pytest.raises(
        SenseAcquiringRequiresCheckboxError,
        match="sense-acquiring.csv.*рядок 2.*Z-звіту Checkbox",
    ):
        run_income_book_pipeline(
            client=_client_profile(),
            bank_statements=[
                BankStatementSource(bank=BankName.SENSE, path=statement_path),
            ],
            checkbox_path=None,
            template_path=template_path,
            output_path=output_path,
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_pipeline_ignores_outgoing_transaction_from_adjacent_month_for_period(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    statement_path = tmp_path / "august-sense.csv"
    checkbox_path = tmp_path / "august-checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    statement_path.write_text("synthetic", encoding="utf-8")
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    credit = BankTransaction(
        source=TransactionSource(
            original_filename=statement_path.name,
            row_number=2,
        ),
        date=date(2026, 6, 1),
        bank=BankName.SENSE,
        account_number="UA273000010000000000000000001",
        currency="UAH",
        document_number="CREDIT-001",
        debit=Decimal("0.00"),
        credit=Decimal("20.00"),
        counterparty="ТОВ Тестовий покупець",
        counterparty_account="UA753000010000000000000000010",
        counterparty_tax_id="11111111",
        payment_purpose="Оплата за послуги",
    )
    debit = credit.model_copy(
        update={
            "date": date(2026, 7, 1),
            "document_number": "DEBIT-001",
            "debit": Decimal("5.00"),
            "credit": Decimal("0.00"),
        }
    )
    monkeypatch.setattr(
        "income_book_automation.pipeline._parse_bank_statement",
        lambda *_args, **_kwargs: [credit, debit],
    )

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[
            BankStatementSource(bank=BankName.SENSE, path=statement_path),
        ],
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert result.output_path == output_path


def test_run_income_book_pipeline_deduplicates_overlapping_statements(
    tmp_path: Path,
) -> None:
    first_statement_path = tmp_path / "first-statement.csv"
    second_statement_path = tmp_path / "second-statement.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_overlapping_pumb_statement(
        first_statement_path,
        unique_document_number="TEST-UNIQUE-001",
        unique_amount="5.00",
    )
    _write_overlapping_pumb_statement(
        second_statement_path,
        unique_document_number="TEST-UNIQUE-002",
        unique_amount="7.00",
    )
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[
            BankStatementSource(
                bank=BankName.PUMB,
                path=first_statement_path,
            ),
            BankStatementSource(
                bank=BankName.PUMB,
                path=second_statement_path,
            ),
        ],
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert len(result.classified_transactions) == 3
    assert result.daily_entries[0].bank_income == Decimal("32.00")
    assert len(result.duplicate_transactions) == 1
    assert result.duplicate_transactions[0].document_number == "TEST-DUPLICATE-001"


def test_run_income_book_pipeline_rejects_duplicate_statement_files(
    tmp_path: Path,
) -> None:
    first_statement_path = tmp_path / "first-statement.csv"
    renamed_copy_path = tmp_path / "renamed-copy.csv"

    _write_pumb_statement(first_statement_path)
    renamed_copy_path.write_bytes(first_statement_path.read_bytes())

    with pytest.raises(
        IncomeBookPipelineError,
        match="renamed-copy.csv.*повторює.*first-statement.csv",
    ):
        run_income_book_pipeline(
            client=_client_profile(),
            bank_statements=[
                BankStatementSource(
                    bank=BankName.PUMB,
                    path=first_statement_path,
                ),
                BankStatementSource(
                    bank=BankName.PUMB,
                    path=renamed_copy_path,
                ),
            ],
            checkbox_path=tmp_path / "checkbox.xlsx",
            template_path=tmp_path / "template.xlsx",
            output_path=tmp_path / "output.xlsx",
            sheet_name="2026",
        )


def test_run_income_book_pipeline_uses_each_mono_account(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    first_statement_path = tmp_path / "first-mono.csv"
    second_statement_path = tmp_path / "second-mono.csv"
    checkbox_path = tmp_path / "checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    first_statement_path.write_text("first", encoding="utf-8")
    second_statement_path.write_text("second", encoding="utf-8")
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    parsed_sources: list[tuple[Path, BankName, str | None]] = []

    def fake_parse_bank_statement(
        path: Path,
        bank: BankName,
        *,
        account_number: str | None,
    ) -> list[BankTransaction]:
        parsed_sources.append((path, bank, account_number))
        return []

    monkeypatch.setattr(
        "income_book_automation.pipeline._parse_bank_statement",
        fake_parse_bank_statement,
    )

    run_income_book_pipeline(
        client=_client_profile(),
        bank_statements=[
            BankStatementSource(
                bank=BankName.MONO,
                path=first_statement_path,
                account_number="UA433000010000000000000000101",
            ),
            BankStatementSource(
                bank=BankName.MONO,
                path=second_statement_path,
                account_number="UA163000010000000000000000102",
            ),
        ],
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert parsed_sources == [
        (
            first_statement_path,
            BankName.MONO,
            "UA433000010000000000000000101",
        ),
        (
            second_statement_path,
            BankName.MONO,
            "UA163000010000000000000000102",
        ),
    ]


def test_run_income_book_pipeline_rejects_non_uah_before_export(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    statement_path = tmp_path / "uploaded-euro-statement.csv"
    output_path = tmp_path / "output.xlsx"
    statement_path.write_text("synthetic statement", encoding="utf-8")

    transaction = BankTransaction(
        source=TransactionSource(
            original_filename="original-euro-statement.csv",
            row_number=7,
        ),
        date=date(2026, 6, 1),
        bank=BankName.PUMB,
        account_number="UA273000010000000000000000001",
        currency="EUR",
        document_number="TEST-EUR-001",
        debit=Decimal("0.00"),
        credit=Decimal("100.00"),
        counterparty="ТОВ Тестовий покупець",
        counterparty_account="UA753000010000000000000000010",
        counterparty_tax_id="11111111",
        payment_purpose="Оплата за тестові послуги",
    )

    monkeypatch.setattr(
        "income_book_automation.pipeline._parse_bank_statement",
        lambda *_args, **_kwargs: [transaction],
    )

    with pytest.raises(
        UnsupportedCurrencyError,
        match=("original-euro-statement.csv.*рядок 7.*EUR.*лише у валюті UAH"),
    ):
        run_income_book_pipeline(
            client=_client_profile(),
            bank=BankName.PUMB,
            bank_statement_path=statement_path,
            checkbox_path=tmp_path / "checkbox.xlsx",
            template_path=tmp_path / "template.xlsx",
            output_path=output_path,
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_pipeline_rejects_bank_and_checkbox_from_different_months(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "june-bank-statement.csv"
    checkbox_path = tmp_path / "july-checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_pumb_statement(statement_path)
    _write_checkbox_report(checkbox_path)
    _write_income_book_template(template_path)

    checkbox_workbook = load_workbook(checkbox_path)
    try:
        checkbox_workbook.active["A2"] = datetime(2026, 7, 1, 12, 0)  # noqa: DTZ001
        checkbox_workbook.save(checkbox_path)
    finally:
        checkbox_workbook.close()

    with pytest.raises(MixedPeriodError) as error_info:
        run_income_book_pipeline(
            client=_client_profile(),
            bank=BankName.PUMB,
            bank_statement_path=statement_path,
            checkbox_path=checkbox_path,
            template_path=template_path,
            output_path=output_path,
            sheet_name="2026",
        )

    error_message = str(error_info.value)
    assert "june-bank-statement.csv" in error_message
    assert "july-checkbox.xlsx" in error_message
    assert "2026-06" in error_message
    assert "2026-07" in error_message
    assert "один календарний місяць" in error_message
    assert not output_path.exists()


def test_pipeline_exports_unchanged_book_with_no_income_warning(
    tmp_path: Path,
) -> None:
    statement_path = tmp_path / "statement-without-income.csv"
    checkbox_path = tmp_path / "zero-checkbox.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    _write_pumb_statement(
        statement_path,
        first_payment_purpose="Повернення коштів покупцю",
    )
    _write_checkbox_report(
        checkbox_path,
        card_revenue=0,
        card_refund=0,
        cash_revenue=0,
        cash_refund=0,
    )
    _write_income_book_template(template_path)

    result = run_income_book_pipeline(
        client=_client_profile(),
        bank=BankName.PUMB,
        bank_statement_path=statement_path,
        checkbox_path=checkbox_path,
        template_path=template_path,
        output_path=output_path,
        sheet_name="2026",
    )

    assert output_path.exists()
    assert result.daily_entries == ()
    assert result.no_income is True
    assert [record.category for record in result.classified_transactions] == [
        TransactionCategory.EXCLUDED,
        TransactionCategory.OWN_TRANSFER,
    ]

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        assert sheet["A6"].value.date() == date(2026, 6, 1)
        assert sheet["B6"].value is None
    finally:
        workbook.close()
