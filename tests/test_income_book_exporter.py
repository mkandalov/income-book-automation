from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from income_book_automation.exporters.income_book import (
    DuplicateMonthTotalRowError,
    HelperColumnMapping,
    IncomeBookExportError,
    IncomeBookTemplateReadError,
    InvalidHelperColumnMappingError,
    MissingIncomeBookDateError,
    MissingIncomeBookSheetError,
    MissingMonthTotalRowError,
    MissingYearTotalRowError,
    export_income_book,
)
from income_book_automation.models import DailyIncomeBookEntry


def _create_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026"

    sheet["A6"] = date(2026, 6, 1)
    sheet["A6"].number_format = "m/d/yy"
    sheet["J6"] = "=K6+L6+M6"
    sheet["J7"] = "=SUM(J6:J6)"
    sheet["K6"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    workbook.save(path)
    workbook.close()


def _entry(
    day: int = 1,
    *,
    month: int = 6,
) -> DailyIncomeBookEntry:
    return DailyIncomeBookEntry(
        date=date(2026, month, day),
        checkbox_card_income=Decimal("90.00"),
        checkbox_cash_income=Decimal("50.00"),
        bank_income=Decimal("20.00"),
    )


def _create_book_through_may(
    path: Path,
    *,
    may_total_label: str = "Всього травень:",
    year_total_label: str = "Всього 2026 рік:",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026"

    sheet["A6"] = date(2026, 4, 1)
    sheet["B6"] = 10
    sheet["D6"] = "=B6-C6"
    sheet["F6"] = "=D6+E6"
    sheet["J6"] = "=K6+L6+M6"
    sheet["A7"] = "Всього квітень:"
    sheet["B7"] = "=SUM(B6:B6)"

    sheet["A8"] = date(2026, 5, 1)
    sheet["B8"] = 20
    sheet["D8"] = "=B8-C8"
    sheet["F8"] = "=D8+E8"
    sheet["J8"] = "=K8+L8+M8"
    sheet["A9"] = may_total_label
    sheet["B9"] = "=SUM(B8:B8)"

    sheet["A10"] = year_total_label
    sheet["B10"] = "=B7+B9"

    sheet["A8"].number_format = "m/d/yy"
    sheet["K8"].fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    sheet["A9"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    workbook.save(path)
    workbook.close()


TOTAL_FORMULA_COLUMNS = ("B", "C", "E", "H", "J", "K", "L", "M")
TEST_MONTH_NAMES = {
    1: "січень",
    2: "лютий",
    3: "березень",
    4: "квітень",
    5: "травень",
    6: "червень",
    7: "липень",
    8: "серпень",
    9: "вересень",
    10: "жовтень",
    11: "листопад",
    12: "грудень",
}


def _write_existing_daily_row(
    sheet: Worksheet,
    row_number: int,
    transaction_date: date,
) -> None:
    sheet.cell(row=row_number, column=1).value = transaction_date

    for column in (2, 3, 5, 8, 10, 11, 12, 13):
        sheet.cell(row=row_number, column=column).value = Decimal(
            row_number * 100 + column
        )

    sheet.cell(row=row_number, column=4).value = f"=B{row_number}-C{row_number}"
    sheet.cell(row=row_number, column=6).value = f"=D{row_number}+E{row_number}"
    sheet.cell(row=row_number, column=7).value = Decimal("0.00")


def _write_existing_month_total(
    sheet: Worksheet,
    row_number: int,
    month_name: str,
    first_daily_row: int,
    last_daily_row: int,
) -> None:
    sheet.cell(row=row_number, column=1).value = f"Всього {month_name}:"

    for column in TOTAL_FORMULA_COLUMNS:
        sheet[f"{column}{row_number}"] = (
            f"=SUM({column}{first_daily_row}:{column}{last_daily_row})"
        )

    sheet.cell(row=row_number, column=4).value = f"=B{row_number}-C{row_number}"
    sheet.cell(row=row_number, column=6).value = f"=D{row_number}+E{row_number}"
    sheet.cell(row=row_number, column=7).value = Decimal("0.00")


def _create_book_with_first_five_months(path: Path) -> dict[int, int]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026"

    month_days = {
        1: ("січень", [2, 19]),
        2: ("лютий", [11]),
        3: ("березень", [1, 14, 27]),
        4: ("квітень", [3, 22]),
        5: ("травень", [2, 9, 18, 30]),
    }
    month_total_rows: dict[int, int] = {}
    row_number = 6

    for month, (month_name, days) in month_days.items():
        first_daily_row = row_number

        for day in days:
            _write_existing_daily_row(
                sheet,
                row_number,
                date(2026, month, day),
            )
            row_number += 1

        month_total_rows[month] = row_number
        _write_existing_month_total(
            sheet,
            row_number,
            month_name,
            first_daily_row,
            row_number - 1,
        )

        if month == 5:
            thin_side = Side(style="thin", color="000000")
            full_border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
            vertical_border = Border(
                left=thin_side,
                right=thin_side,
            )

            for column in range(1, 9):
                sheet.cell(row=row_number, column=column).border = full_border

            sheet.cell(row=row_number, column=10).border = full_border

            for column in range(11, 15):
                sheet.cell(row=row_number, column=column).border = vertical_border

        row_number += 1

        if month == 3:
            sheet.cell(row=row_number, column=1).value = "Всього 1 кв 2026:"

            for column in TOTAL_FORMULA_COLUMNS:
                references = "+".join(
                    f"{column}{month_total_rows[quarter_month]}"
                    for quarter_month in range(1, 4)
                )
                sheet[f"{column}{row_number}"] = f"={references}"

            row_number += 1

    sheet.cell(row=row_number, column=1).value = "Всього 2026 рік:"
    stale_cell = sheet.cell(row=row_number, column=9)
    stale_cell.value = "stale value"
    stale_cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    stale_cell.border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    sheet.cell(row=row_number, column=14).value = "=N8+N10+N14+N18+N23"

    for column in TOTAL_FORMULA_COLUMNS:
        references = "+".join(
            f"{column}{month_total_rows[existing_month]}"
            for existing_month in range(1, 6)
        )
        sheet[f"{column}{row_number}"] = f"={references}"

    workbook.save(path)
    workbook.close()
    return month_total_rows


def _create_book_before_quarter_end(
    path: Path,
    ending_month: int,
) -> tuple[dict[int, int], int]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026"
    month_total_rows: dict[int, int] = {}
    row_number = 6

    for month in range(1, ending_month):
        first_daily_row = row_number
        day_count = month % 3 + 1

        for day in range(1, day_count + 1):
            _write_existing_daily_row(
                sheet,
                row_number,
                date(2026, month, day),
            )
            row_number += 1

        month_total_rows[month] = row_number
        _write_existing_month_total(
            sheet,
            row_number,
            TEST_MONTH_NAMES[month],
            first_daily_row,
            row_number - 1,
        )
        row_number += 1

        if month % 3 == 0:
            quarter = (month - 1) // 3 + 1
            quarter_months = range(month - 2, month + 1)
            sheet.cell(row=row_number, column=1).value = f"Всього {quarter} кв 2026:"

            for column in TOTAL_FORMULA_COLUMNS:
                references = "+".join(
                    f"{column}{month_total_rows[quarter_month]}"
                    for quarter_month in quarter_months
                )
                sheet[f"{column}{row_number}"] = f"={references}"

            row_number += 1

        if month == 6:
            sheet.cell(row=row_number, column=1).value = "Всього 1 півріччя 2026:"
            row_number += 1

    year_total_row = row_number
    sheet.cell(row=year_total_row, column=1).value = "Всього 2026 рік:"

    for column in TOTAL_FORMULA_COLUMNS:
        references = "+".join(
            f"{column}{month_total_rows[month]}" for month in range(1, ending_month)
        )
        sheet[f"{column}{year_total_row}"] = f"={references}"

    workbook.save(path)
    workbook.close()
    return month_total_rows, year_total_row


def test_export_income_book_updates_matching_date_and_preserves_template(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output" / "income-book.xlsx"
    _create_template(template_path)
    original_template = template_path.read_bytes()

    result = export_income_book(
        template_path,
        output_path,
        [_entry()],
        sheet_name="2026",
    )

    assert result == output_path
    assert template_path.read_bytes() == original_template

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        assert sheet["A6"].value.date() == date(2026, 6, 1)
        assert sheet["B6"].value == 160
        assert sheet["C6"].value == 0
        assert sheet["D6"].value == "=B6-C6"
        assert sheet["E6"].value == 0
        assert sheet["F6"].value == "=D6+E6"
        assert sheet["G6"].value == 0
        assert sheet["H6"].value == 0
        assert sheet["K6"].value == 90
        assert sheet["L6"].value == 50
        assert sheet["M6"].value == 20
        assert sheet["J6"].value == "=K6+L6+M6"
        assert sheet["J7"].value == "=SUM(J6:J6)"
        assert sheet["K6"].fill.fgColor.rgb.endswith("FFFF00")
    finally:
        workbook.close()


def test_export_income_book_supports_custom_helper_columns(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_template(template_path)

    helper_columns = HelperColumnMapping(
        total=10,
        checkbox_card=12,
        checkbox_cash=13,
        bank_income=14,
    )

    export_income_book(
        template_path,
        output_path,
        [_entry()],
        sheet_name="2026",
        helper_columns=helper_columns,
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]

        assert sheet["J6"].value == "=L6+M6+N6"
        assert sheet["K6"].value is None
        assert sheet["L6"].value == 90
        assert sheet["M6"].value == 50
        assert sheet["N6"].value == 20
    finally:
        workbook.close()


def test_helper_column_mapping_rejects_duplicate_columns() -> None:
    with pytest.raises(
        InvalidHelperColumnMappingError,
        match="Кожен показник",
    ):
        HelperColumnMapping(
            total=10,
            checkbox_card=11,
            checkbox_cash=11,
            bank_income=13,
        )


def test_helper_column_mapping_rejects_official_columns() -> None:
    with pytest.raises(
        InvalidHelperColumnMappingError,
        match=r"починатися з колонки 10 \(J\)",
    ):
        HelperColumnMapping(
            total=9,
            checkbox_card=11,
            checkbox_cash=12,
            bank_income=13,
        )


def test_helper_column_mapping_rejects_columns_after_fifteen() -> None:
    with pytest.raises(
        InvalidHelperColumnMappingError,
        match=r"правіше колонки 15 \(O\)",
    ):
        HelperColumnMapping(
            total=10,
            checkbox_card=11,
            checkbox_cash=12,
            bank_income=16,
        )


def test_export_income_book_rejects_missing_date(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_template(template_path)

    with pytest.raises(MissingIncomeBookDateError, match="2026-06-02"):
        export_income_book(
            template_path,
            output_path,
            [_entry(day=2)],
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_export_income_book_rejects_missing_sheet(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_template(template_path)

    with pytest.raises(MissingIncomeBookSheetError, match="missing-sheet"):
        export_income_book(
            template_path,
            output_path,
            [_entry()],
            sheet_name="missing-sheet",
        )

    assert not output_path.exists()


def test_export_income_book_rejects_unreadable_template(tmp_path: Path) -> None:
    template_path = tmp_path / "broken.xlsx"
    output_path = tmp_path / "output.xlsx"
    template_path.write_bytes(b"not an XLSX workbook")

    with pytest.raises(IncomeBookTemplateReadError, match="broken.xlsx"):
        export_income_book(
            template_path,
            output_path,
            [_entry()],
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_export_income_book_refuses_to_overwrite_template(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _create_template(template_path)
    original_template = template_path.read_bytes()

    with pytest.raises(IncomeBookExportError, match="must differ"):
        export_income_book(
            template_path,
            template_path,
            [_entry()],
            sheet_name="2026",
        )

    assert template_path.read_bytes() == original_template


def test_export_income_book_appends_new_month_and_period_totals(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(template_path)

    export_income_book(
        template_path,
        output_path,
        [_entry(day=3), _entry(day=1)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]

        assert sheet["A10"].value.date() == date(2026, 6, 1)
        assert sheet["A11"].value.date() == date(2026, 6, 3)
        assert sheet["B10"].value == 160
        assert sheet["B11"].value == 160
        assert sheet["D10"].value == "=B10-C10"
        assert sheet["F10"].value == "=D10+E10"
        assert sheet["J10"].value == "=K10+L10+M10"
        assert sheet["K10"].fill.fgColor.rgb.endswith("D9EAD3")

        assert sheet["A12"].value == "Всього червень:"
        assert sheet["B12"].value == "=SUM(B10:B11)"
        assert sheet["D12"].value == "=B12-C12"
        assert sheet["F12"].value == "=D12+E12"
        assert sheet["A12"].fill.fgColor.rgb.endswith("FFFF00")

        assert sheet["A13"].value == "Всього 2 кв 2026:"
        assert sheet["B13"].value == "=B7+B9+B12"

        assert sheet["A14"].value == "Всього 1 півріччя 2026:"
        assert sheet["B14"].value == "=B7+B9+B12"

        assert sheet["A15"].value == "Всього 2026 рік:"
        assert sheet["B15"].value == "=B7+B9+B12"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "year_total_label",
    [
        "Всього 2026 рік:",
        "Всього 2026 рік",
        "Всього 2026 рік:   ",
        "Всього\xa02026\xa0рік:",
        "ВСЬОГО 2026 РІК:",
    ],
)
def test_export_income_book_accepts_year_total_label_variants(
    tmp_path: Path,
    year_total_label: str,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(
        template_path,
        year_total_label=year_total_label,
    )

    export_income_book(
        template_path,
        output_path,
        [_entry(day=1)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook["2026"]["A14"].value == "Всього 2026 рік:"
    finally:
        workbook.close()


def test_export_income_book_rejects_missing_year_total_row(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(
        template_path,
        year_total_label="Нерозпізнаний річний підсумок",
    )

    with pytest.raises(
        MissingYearTotalRowError,
        match="Всього 2026 рік:",
    ):
        export_income_book(
            template_path,
            output_path,
            [_entry(day=1)],
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_export_income_book_repairs_misspelled_month_total_label(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(
        template_path,
        may_total_label="Всььго травень:",
    )

    export_income_book(
        template_path,
        output_path,
        [_entry(day=3), _entry(day=1)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]

        assert sheet["B13"].value == "=B7+B9+B12"
        assert sheet["B14"].value == "=B7+B9+B12"
        assert sheet["B15"].value == "=B7+B9+B12"
        assert sheet["A9"].value == "Всього травень:"
    finally:
        workbook.close()


def test_export_income_book_rejects_duplicate_month_total_rows(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(template_path)

    workbook = load_workbook(template_path)
    try:
        sheet = workbook["2026"]
        sheet.insert_rows(10)
        sheet["A10"] = "Всього травень:"
        sheet["B10"] = "=SUM(B8:B8)"
        workbook.save(template_path)
    finally:
        workbook.close()

    with pytest.raises(
        DuplicateMonthTotalRowError,
        match="multiple total rows found for month: травень",
    ):
        export_income_book(
            template_path,
            output_path,
            [_entry(day=1), _entry(day=3)],
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_export_income_book_rejects_month_data_without_total_row(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(
        template_path,
        may_total_label="Нерозпізнаний підсумок:",
    )

    with pytest.raises(
        MissingMonthTotalRowError,
        match="month total row not found for existing data: травень",
    ):
        export_income_book(
            template_path,
            output_path,
            [_entry(day=1), _entry(day=3)],
            sheet_name="2026",
        )

    assert not output_path.exists()


def test_export_income_book_uses_actual_month_totals_for_all_periods(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    month_total_rows = _create_book_with_first_five_months(template_path)

    export_income_book(
        template_path,
        output_path,
        [_entry(day=21), _entry(day=1), _entry(day=7)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        june_daily_rows = [24, 25, 26]
        june_total_row = 27
        second_quarter_row = 28
        half_year_row = 29
        year_total_row = 30

        assert [
            sheet.cell(row=row_number, column=1).value.date()
            for row_number in june_daily_rows
        ] == [
            date(2026, 6, 1),
            date(2026, 6, 7),
            date(2026, 6, 21),
        ]

        month_total_rows[6] = june_total_row
        second_quarter_months = range(4, 7)
        half_year_months = range(1, 7)

        for column in TOTAL_FORMULA_COLUMNS:
            assert sheet[f"{column}{june_total_row}"].value == (
                f"=SUM({column}24:{column}26)"
            )

            second_quarter_references = "+".join(
                f"{column}{month_total_rows[month]}" for month in second_quarter_months
            )
            assert sheet[f"{column}{second_quarter_row}"].value == (
                f"={second_quarter_references}"
            )

            half_year_references = "+".join(
                f"{column}{month_total_rows[month]}" for month in half_year_months
            )
            assert sheet[f"{column}{half_year_row}"].value == (
                f"={half_year_references}"
            )
            assert sheet[f"{column}{year_total_row}"].value == (
                f"={half_year_references}"
            )

        assert sheet[f"D{second_quarter_row}"].value == (
            f"=B{second_quarter_row}-C{second_quarter_row}"
        )
        assert sheet[f"F{half_year_row}"].value == (
            f"=D{half_year_row}+E{half_year_row}"
        )
        assert sheet["B6"].value == 602
    finally:
        workbook.close()


def test_export_income_book_clears_unused_year_total_column(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_with_first_five_months(template_path)

    export_income_book(
        template_path,
        output_path,
        [_entry(day=1), _entry(day=7), _entry(day=21)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        cell = sheet["I30"]
        assert cell.value is None
        assert cell.fill.fill_type is None
        assert cell.border.left.style is None
        assert cell.border.right.style is None
        assert cell.border.top.style is None
        assert cell.border.bottom.style is None
    finally:
        workbook.close()


def test_export_income_book_adds_grid_to_generated_helper_totals(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_with_first_five_months(template_path)

    export_income_book(
        template_path,
        output_path,
        [_entry(day=1), _entry(day=7), _entry(day=21)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]

        for row_number in (27, 28, 29):
            for column in range(10, 14):
                border = sheet.cell(row=row_number, column=column).border
                assert border.left.style == "thin"
                assert border.right.style == "thin"
                assert getattr(border.top, "style", None) == "thin"
                assert getattr(border.bottom, "style", None) == "thin"
    finally:
        workbook.close()


def test_export_income_book_clears_unselected_helper_column(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_with_first_five_months(template_path)

    export_income_book(
        template_path,
        output_path,
        [_entry(day=1), _entry(day=7), _entry(day=21)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        assert sheet["N30"].value is None
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("ending_month", "quarter"),
    [
        (3, 1),
        (6, 2),
        (9, 3),
        (12, 4),
    ],
)
def test_export_income_book_builds_each_quarter_from_its_three_months(
    tmp_path: Path,
    ending_month: int,
    quarter: int,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    month_total_rows, original_year_total_row = _create_book_before_quarter_end(
        template_path,
        ending_month,
    )

    export_income_book(
        template_path,
        output_path,
        [_entry(day=17, month=ending_month)],
        sheet_name="2026",
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook["2026"]
        new_month_total_row = original_year_total_row + 1
        quarter_total_row = original_year_total_row + 2
        inserted_rows = 4 if ending_month == 6 else 3
        shifted_year_total_row = original_year_total_row + inserted_rows
        month_total_rows[ending_month] = new_month_total_row
        quarter_months = range(ending_month - 2, ending_month + 1)

        assert sheet.cell(row=quarter_total_row, column=1).value == (
            f"Всього {quarter} кв 2026:"
        )

        for column in TOTAL_FORMULA_COLUMNS:
            quarter_references = "+".join(
                f"{column}{month_total_rows[month]}" for month in quarter_months
            )
            assert sheet[f"{column}{quarter_total_row}"].value == (
                f"={quarter_references}"
            )

            year_references = "+".join(
                f"{column}{month_total_rows[month]}"
                for month in range(1, ending_month + 1)
            )
            assert sheet[f"{column}{shifted_year_total_row}"].value == (
                f"={year_references}"
            )

        assert sheet.cell(row=quarter_total_row, column=9).value is None
        assert sheet.cell(row=shifted_year_total_row, column=9).value is None
    finally:
        workbook.close()


def test_export_income_book_rejects_entries_from_multiple_months(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _create_book_through_may(template_path)

    entries = [
        _entry(day=1),
        DailyIncomeBookEntry(
            date=date(2026, 7, 1),
            checkbox_card_income=Decimal("10.00"),
            checkbox_cash_income=Decimal("0.00"),
            bank_income=Decimal("0.00"),
        ),
    ]

    with pytest.raises(IncomeBookExportError, match="one calendar month"):
        export_income_book(
            template_path,
            output_path,
            entries,
            sheet_name="2026",
        )

    assert not output_path.exists()
