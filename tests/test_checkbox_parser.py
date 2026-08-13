from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from income_book_automation.parsers.checkbox import (
    InvalidCheckboxRowError,
    MissingCheckboxColumnError,
    parse_checkbox_file,
    parse_checkbox_row,
)

COLUMN_COUNT = 29
COLUMN_INDEXES = {
    "opened_at": 1,
    "card_revenue": 25,
    "card_refund": 26,
    "cash_revenue": 27,
    "cash_refund": 28,
}


def _make_checkbox_headers() -> list[object]:
    headers: list[object] = [None] * COLUMN_COUNT
    headers[1] = "Дата відкриття"
    headers[25] = "Виручка безготівка"
    headers[26] = "Повернення безготівка"
    headers[27] = "Виручка готівка"
    headers[28] = "Повернення готівка"
    return headers


def _excel_datetime(
    year: int,
    month: int,
    day: int,
    hour: int,
    minute: int,
) -> datetime:
    # Excel stores timestamps without timezone information.
    return datetime(year, month, day, hour, minute)  # noqa: DTZ001


def _make_checkbox_row(
    opened_at: datetime | None,
    *,
    card_revenue: object,
    card_refund: object,
    cash_revenue: object,
    cash_refund: object,
) -> tuple[object, ...]:
    row: list[object] = [None] * COLUMN_COUNT
    row[1] = opened_at
    row[25] = card_revenue
    row[26] = card_refund
    row[27] = cash_revenue
    row[28] = cash_refund
    return tuple(row)


def test_parse_checkbox_row_maps_values_and_calculates_net() -> None:
    row = _make_checkbox_row(
        _excel_datetime(2026, 6, 18, 8, 49),
        card_revenue=37331,
        card_refund=844,
        cash_revenue=1000,
        cash_refund=100,
    )

    record = parse_checkbox_row(
        row,
        COLUMN_INDEXES,
        Path("test-checkbox.xlsx"),
        2,
    )

    assert record.date == date(2026, 6, 18)
    assert record.card_net == Decimal(36487)
    assert record.cash_net == Decimal(900)
    assert record.total_net == Decimal(37387)


def test_parse_checkbox_row_accepts_explicit_zero_amounts() -> None:
    row = _make_checkbox_row(
        _excel_datetime(2026, 6, 19, 9, 0),
        card_revenue=0,
        card_refund="0.00",
        cash_revenue=Decimal("0.00"),
        cash_refund=0.0,
    )

    record = parse_checkbox_row(
        row,
        COLUMN_INDEXES,
        Path("test-checkbox.xlsx"),
        2,
    )

    assert record.card_revenue == Decimal(0)
    assert record.card_refund == Decimal(0)
    assert record.cash_revenue == Decimal(0)
    assert record.cash_refund == Decimal(0)
    assert record.total_net == Decimal(0)


def test_parse_checkbox_row_accepts_refund_equal_to_revenue() -> None:
    row = _make_checkbox_row(
        _excel_datetime(2026, 6, 19, 9, 0),
        card_revenue=100,
        card_refund=100,
        cash_revenue=50,
        cash_refund=50,
    )

    record = parse_checkbox_row(
        row,
        COLUMN_INDEXES,
        Path("test-checkbox.xlsx"),
        2,
    )

    assert record.card_net == Decimal(0)
    assert record.cash_net == Decimal(0)
    assert record.total_net == Decimal(0)


@pytest.mark.parametrize(
    ("overrides", "expected_card_net", "expected_cash_net"),
    [
        (
            {"card_revenue": 100, "card_refund": 101},
            Decimal(-1),
            Decimal(50),
        ),
        (
            {"cash_revenue": 50, "cash_refund": 51},
            Decimal(100),
            Decimal(-1),
        ),
    ],
)
def test_parse_checkbox_row_accepts_refund_greater_than_revenue(
    overrides: dict[str, object],
    expected_card_net: Decimal,
    expected_cash_net: Decimal,
) -> None:
    values: dict[str, object] = {
        "card_revenue": 100,
        "card_refund": 0,
        "cash_revenue": 50,
        "cash_refund": 0,
    }
    values.update(overrides)

    row = _make_checkbox_row(
        _excel_datetime(2026, 6, 19, 9, 0),
        card_revenue=values["card_revenue"],
        card_refund=values["card_refund"],
        cash_revenue=values["cash_revenue"],
        cash_refund=values["cash_refund"],
    )

    record = parse_checkbox_row(
        row,
        COLUMN_INDEXES,
        Path("test-checkbox.xlsx"),
        7,
    )

    assert record.card_net == expected_card_net
    assert record.cash_net == expected_cash_net


@pytest.mark.parametrize(
    ("column_key", "column_name", "missing_value"),
    [
        ("card_revenue", "Виручка безготівка", None),
        ("card_revenue", "Виручка безготівка", "   "),
        ("card_refund", "Повернення безготівка", None),
        ("card_refund", "Повернення безготівка", ""),
        ("cash_revenue", "Виручка готівка", None),
        ("cash_revenue", "Виручка готівка", "\t"),
        ("cash_refund", "Повернення готівка", None),
        ("cash_refund", "Повернення готівка", "\n"),
    ],
)
def test_parse_checkbox_row_rejects_each_missing_amount(
    column_key: str,
    column_name: str,
    missing_value: object,
) -> None:
    values: dict[str, object] = {
        "card_revenue": 100,
        "card_refund": 0,
        "cash_revenue": 50,
        "cash_refund": 0,
    }
    values[column_key] = missing_value

    row = _make_checkbox_row(
        _excel_datetime(2026, 6, 19, 9, 0),
        card_revenue=values["card_revenue"],
        card_refund=values["card_refund"],
        cash_revenue=values["cash_revenue"],
        cash_refund=values["cash_refund"],
    )

    with pytest.raises(
        InvalidCheckboxRowError,
        match=rf"column '{column_name}': required value is missing",
    ):
        parse_checkbox_row(
            row,
            COLUMN_INDEXES,
            Path("test-checkbox.xlsx"),
            7,
        )


def test_parse_checkbox_row_rejects_missing_date() -> None:
    row = _make_checkbox_row(
        None,
        card_revenue=100,
        card_refund=0,
        cash_revenue=50,
        cash_refund=0,
    )

    with pytest.raises(
        InvalidCheckboxRowError,
        match="column 'Дата відкриття': required value is missing",
    ):
        parse_checkbox_row(
            row,
            COLUMN_INDEXES,
            Path("test-checkbox.xlsx"),
            7,
        )


def test_parse_checkbox_file_reads_rows_and_skips_empty_date(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(_make_checkbox_headers())
    worksheet.append(
        list(
            _make_checkbox_row(
                _excel_datetime(2026, 6, 1, 8, 0),
                card_revenue=1000,
                card_refund=100,
                cash_revenue=200,
                cash_refund=50,
            )
        )
    )
    worksheet.append([None] * COLUMN_COUNT)
    worksheet.append(
        list(
            _make_checkbox_row(
                _excel_datetime(2026, 6, 2, 8, 0),
                card_revenue=2000,
                card_refund=0,
                cash_revenue=300,
                cash_refund=0,
            )
        )
    )

    source_path = tmp_path / "checkbox.xlsx"
    workbook.save(source_path)
    workbook.close()

    records = parse_checkbox_file(source_path)

    assert len(records) == 2
    assert records[0].date == date(2026, 6, 1)
    assert records[0].total_net == Decimal(1050)
    assert records[1].date == date(2026, 6, 2)
    assert records[1].total_net == Decimal(2300)


def test_parse_checkbox_file_rejects_undated_row_with_amounts(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(_make_checkbox_headers())
    worksheet.append(
        list(
            _make_checkbox_row(
                None,
                card_revenue=100,
                card_refund=0,
                cash_revenue=50,
                cash_refund=0,
            )
        )
    )

    source_path = tmp_path / "undated-checkbox-row.xlsx"
    workbook.save(source_path)
    workbook.close()

    with pytest.raises(
        InvalidCheckboxRowError,
        match="row 2, column 'Дата відкриття': required value is missing",
    ):
        parse_checkbox_file(source_path)


def test_parse_checkbox_file_finds_reordered_columns(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "Виручка готівка",
            "Повернення безготівка",
            "Дата відкриття",
            "Повернення готівка",
            "Виручка безготівка",
        ]
    )
    worksheet.append(
        [
            500,
            100,
            _excel_datetime(2026, 6, 18, 8, 49),
            50,
            1000,
        ]
    )

    source_path = tmp_path / "reordered-checkbox.xlsx"
    workbook.save(source_path)
    workbook.close()

    records = parse_checkbox_file(source_path)

    assert len(records) == 1
    assert records[0].date == date(2026, 6, 18)
    assert records[0].card_net == Decimal(900)
    assert records[0].cash_net == Decimal(450)
    assert records[0].total_net == Decimal(1350)


def test_parse_checkbox_file_rejects_formula_without_cached_result(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(_make_checkbox_headers())

    row = list(
        _make_checkbox_row(
            _excel_datetime(2026, 6, 18, 8, 49),
            card_revenue=1000,
            card_refund=100,
            cash_revenue=500,
            cash_refund=50,
        )
    )
    row[COLUMN_INDEXES["card_revenue"]] = "=500+500"
    worksheet.append(row)

    source_path = tmp_path / "formula-without-cache-checkbox.xlsx"
    workbook.save(source_path)
    workbook.close()

    with pytest.raises(
        InvalidCheckboxRowError,
        match=("row 2, column 'Виручка безготівка': formula has no cached result"),
    ):
        parse_checkbox_file(source_path)


def test_parse_checkbox_file_rejects_missing_required_header(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = _make_checkbox_headers()
    headers[28] = None
    worksheet.append(headers)

    source_path = tmp_path / "missing-header-checkbox.xlsx"
    workbook.save(source_path)
    workbook.close()

    with pytest.raises(MissingCheckboxColumnError) as error_info:
        parse_checkbox_file(source_path)

    assert error_info.value.filename == "missing-header-checkbox.xlsx"
    assert error_info.value.missing_headers == ("Повернення готівка",)
    assert str(error_info.value) == (
        "У Z-звіті Checkbox «missing-header-checkbox.xlsx» відсутня "
        "обов’язкова колонка: «Повернення готівка». "
        "Книгу доходів не сформовано."
    )


def test_parse_checkbox_file_rejects_negative_amount(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(_make_checkbox_headers())
    worksheet.append(
        list(
            _make_checkbox_row(
                _excel_datetime(2026, 6, 18, 8, 49),
                card_revenue=-1,
                card_refund=0,
                cash_revenue=0,
                cash_refund=0,
            )
        )
    )

    source_path = tmp_path / "negative-amount-checkbox.xlsx"
    workbook.save(source_path)
    workbook.close()

    with pytest.raises(InvalidCheckboxRowError, match="invalid monetary values"):
        parse_checkbox_file(source_path)
