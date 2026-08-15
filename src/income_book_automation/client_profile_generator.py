"""Generate private client-profile YAML files from an Excel register."""

from __future__ import annotations

import argparse
import hashlib
import re
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from pydantic import ValidationError

from income_book_automation.models import ClientProfile

NAME_HEADERS = frozenset({"ПІБ", "ФІО", "ФИО"})
TAX_ID_HEADERS = frozenset({"ІПН", "ИНН", "РНОКПП"})
HEADER_SEARCH_LIMIT = 20


class ClientProfileGenerationError(Exception):
    """Raised when the source register cannot be converted safely."""


@dataclass(frozen=True)
class SourceClient:
    source_row: int
    full_name: str
    tax_id: str


def _normalized_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[^A-ZА-ЯІЇЄҐ0-9]", "", text.upper())


def _normalized_name(value: Any, *, row_number: int) -> str:
    if value is None:
        raise ClientProfileGenerationError(f"Row {row_number}: the full name is empty")

    name = " ".join(str(value).replace("\u00a0", " ").split())
    name = re.sub(r"^ФОП\s+", "", name, flags=re.IGNORECASE)
    parts = name.split()

    if len(parts) < 2:
        raise ClientProfileGenerationError(
            f"Row {row_number}: expected at least a surname and first name"
        )

    return name


def _normalized_tax_id(cell: Cell, *, row_number: int) -> str:
    value = cell.value

    if value is None or isinstance(value, bool):
        raise ClientProfileGenerationError(f"Row {row_number}: the tax ID is empty")

    if isinstance(value, int):
        tax_id = str(value)
    elif isinstance(value, float) and value.is_integer():
        tax_id = str(int(value))
    else:
        tax_id = re.sub(r"[\s\u00a0]", "", str(value))

    if len(tax_id) < 10 and _cell_format_preserves_leading_zero(cell):
        tax_id = tax_id.zfill(10)

    if not re.fullmatch(r"\d{10}", tax_id):
        raise ClientProfileGenerationError(
            f"Row {row_number}: the tax ID must contain exactly 10 digits"
        )

    return tax_id


def _cell_format_preserves_leading_zero(cell: Cell) -> bool:
    number_format = cell.number_format.replace(" ", "")
    return re.fullmatch(r"0{10}", number_format) is not None


def _find_columns(sheet: Any) -> tuple[int, int, int]:
    for row in sheet.iter_rows(min_row=1, max_row=HEADER_SEARCH_LIMIT):
        headers = {
            _normalized_header(cell.value): cell.column
            for cell in row
            if cell.value is not None
        }

        name_column = next(
            (headers[header] for header in NAME_HEADERS if header in headers),
            None,
        )
        tax_id_column = next(
            (headers[header] for header in TAX_ID_HEADERS if header in headers),
            None,
        )

        if name_column is not None and tax_id_column is not None:
            return row[0].row, name_column, tax_id_column

    raise ClientProfileGenerationError(
        "Could not find the full-name and tax-ID columns in the first 20 rows"
    )


def read_clients(source_path: Path) -> list[SourceClient]:
    """Read and validate all client rows without creating output files."""
    try:
        workbook = load_workbook(source_path, read_only=False, data_only=True)
    except (OSError, ValueError) as exc:
        raise ClientProfileGenerationError(
            f"Could not open the Excel register: {source_path.name}"
        ) from exc

    try:
        sheet = workbook.active
        header_row, name_column, tax_id_column = _find_columns(sheet)
        clients: list[SourceClient] = []

        for row_number in range(header_row + 1, sheet.max_row + 1):
            name_cell = sheet.cell(row=row_number, column=name_column)
            tax_id_cell = sheet.cell(row=row_number, column=tax_id_column)

            if name_cell.value is None and tax_id_cell.value is None:
                continue

            clients.append(
                SourceClient(
                    source_row=row_number,
                    full_name=_normalized_name(
                        name_cell.value,
                        row_number=row_number,
                    ),
                    tax_id=_normalized_tax_id(
                        tax_id_cell,
                        row_number=row_number,
                    ),
                )
            )
    finally:
        workbook.close()

    if not clients:
        raise ClientProfileGenerationError("The Excel register contains no clients")

    _validate_unique_tax_ids(clients)
    return clients


def _validate_unique_tax_ids(clients: Sequence[SourceClient]) -> None:
    seen: dict[str, int] = {}

    for client in clients:
        previous_row = seen.get(client.tax_id)
        if previous_row is not None:
            raise ClientProfileGenerationError(
                f"Duplicate tax ID in rows {previous_row} and {client.source_row}"
            )
        seen[client.tax_id] = client.source_row


def build_name_aliases(full_name: str) -> list[str]:
    """Return common bank-statement variants of a Ukrainian full name."""
    surname, first_name, *remaining_parts = full_name.split()
    initials = f"{first_name[0]}."
    initials_with_space = initials

    if remaining_parts:
        patronymic = remaining_parts[0]
        initials = f"{initials}{patronymic[0]}."
        initials_with_space = f"{first_name[0]}. {patronymic[0]}."

    return list(
        dict.fromkeys(
            [
                full_name,
                f"ФОП {full_name}",
                f"{surname} {initials}",
                f"{surname} {initials_with_space}",
                f"ФОП {surname} {initials}",
                f"ФОП {surname} {initials_with_space}",
            ]
        )
    )


def build_profile(client: SourceClient) -> dict[str, Any]:
    private_identifier = hashlib.sha256(client.tax_id.encode("ascii")).hexdigest()[:16]
    profile = {
        "client_id": f"client-{private_identifier}",
        "legal_name": f"ФОП {client.full_name}",
        "tax_id": client.tax_id,
        "own_accounts": [],
        "name_aliases": build_name_aliases(client.full_name),
    }

    try:
        ClientProfile.model_validate(profile)
    except ValidationError as exc:
        raise ClientProfileGenerationError(
            f"Row {client.source_row}: generated profile is invalid"
        ) from exc

    return profile


def _safe_filename_part(value: str) -> str:
    filename_part = re.sub(r"[^\w.-]+", "_", value, flags=re.UNICODE)
    return filename_part.strip("._")


def profile_filename(client: SourceClient, position: int) -> str:
    surname, first_name, *_ = client.full_name.split()
    return (
        f"{position:03d}_"
        f"{_safe_filename_part(surname)}_{_safe_filename_part(first_name)}.yaml"
    )


def generate_client_profiles(
    source_path: Path,
    output_directory: Path,
    *,
    expected_count: int | None = None,
    dry_run: bool = False,
) -> int:
    """Validate the register and create one YAML file per client."""
    clients = read_clients(source_path)

    if expected_count is not None and len(clients) != expected_count:
        raise ClientProfileGenerationError(
            f"Expected {expected_count} clients, but found {len(clients)}"
        )

    profiles = [build_profile(client) for client in clients]
    filenames = [
        profile_filename(client, position)
        for position, client in enumerate(clients, start=1)
    ]

    if len(set(filenames)) != len(filenames):
        raise ClientProfileGenerationError("Generated filenames are not unique")

    if dry_run:
        return len(profiles)

    output_directory.mkdir(parents=True, exist_ok=True)
    existing_profiles = list(output_directory.glob("*.yaml"))
    if existing_profiles:
        raise ClientProfileGenerationError(
            "The output directory already contains YAML files; "
            "choose an empty directory"
        )

    for filename, profile in zip(filenames, profiles, strict=True):
        output_path = output_directory / filename
        output_path.write_text(
            yaml.safe_dump(
                profile,
                allow_unicode=True,
                sort_keys=False,
                default_flow_style=False,
            ),
            encoding="utf-8",
        )

    return len(profiles)


def _build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate private client YAML profiles from an Excel register."
    )
    parser.add_argument("--input", type=Path, required=True, help="Source XLSX file")
    parser.add_argument(
        "--output",
        type=Path,
        required=True,
        help="Empty directory for generated YAML files",
    )
    parser.add_argument(
        "--expected-count",
        type=int,
        default=None,
        help="Fail unless the register contains exactly this many clients",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate everything without creating files",
    )
    return parser


def main() -> None:
    args = _build_argument_parser().parse_args()

    try:
        count = generate_client_profiles(
            args.input,
            args.output,
            expected_count=args.expected_count,
            dry_run=args.dry_run,
        )
    except ClientProfileGenerationError as exc:
        raise SystemExit(f"Error: {exc}") from exc

    action = "Validated" if args.dry_run else "Created"
    print(f"{action} {count} client profiles.")
