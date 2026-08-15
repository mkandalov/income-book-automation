"""Exporter for the income-book Excel workbook."""

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from income_book_automation.models import DailyIncomeBookEntry

DATE_COLUMN = 1
INCOME_COLUMN = 2
REFUND_COLUMN = 3
ADJUSTED_INCOME_COLUMN = 4
FREE_GOODS_COLUMN = 5
TOTAL_INCOME_COLUMN = 6
SPECIAL_INCOME_TYPE_COLUMN = 7
SPECIAL_INCOME_AMOUNT_COLUMN = 8
RESERVED_BLANK_COLUMN = 9
MIN_HELPER_COLUMN = 10
MAX_HELPER_COLUMN = 15


MONTH_NAMES_UKR = {
    1: "січень",
    2: "лютий",
    3: "березень",
    4: "квітень",
    5: "травень",
    6: "червень",
    7: "липень",
    8: "серпень",
    9: "вересень",
    10: "жовтень",
    11: "листопад",
    12: "грудень",
}

OFFICIAL_TOTALLED_COLUMNS = (
    INCOME_COLUMN,
    REFUND_COLUMN,
    FREE_GOODS_COLUMN,
    SPECIAL_INCOME_AMOUNT_COLUMN,
)


class IncomeBookExportError(Exception):
    """Base error raised while exporting an income book."""


class InvalidHelperColumnMappingError(IncomeBookExportError):
    """Raised when helper Excel columns are configured incorrectly."""


class MissingIncomeBookSheetError(IncomeBookExportError):
    """Raised when the requested worksheet does not exist."""


class MissingIncomeBookDateError(IncomeBookExportError):
    """Raised when an entry date is absent from the template."""


@dataclass(frozen=True, slots=True)
class HelperColumnMapping:
    total: int = 10
    checkbox_card: int = 11
    checkbox_cash: int = 12
    bank_income: int = 13

    def __post_init__(self) -> None:
        columns = (
            self.total,
            self.checkbox_card,
            self.checkbox_cash,
            self.bank_income,
        )

        if any(column < MIN_HELPER_COLUMN for column in columns):
            raise InvalidHelperColumnMappingError(
                "helper columns must start from column 10"
            )

        if any(column > MAX_HELPER_COLUMN for column in columns):
            raise InvalidHelperColumnMappingError(
                "helper columns must not exceed column 15"
            )

        if len(set(columns)) != len(columns):
            raise InvalidHelperColumnMappingError("helper columns must be unique")


def _totalled_columns(
    helper_columns: HelperColumnMapping,
) -> tuple[int, ...]:
    return (
        *OFFICIAL_TOTALLED_COLUMNS,
        helper_columns.total,
        helper_columns.checkbox_card,
        helper_columns.checkbox_cash,
        helper_columns.bank_income,
    )


def _selected_helper_columns(helper_columns: HelperColumnMapping) -> set[int]:
    return {
        helper_columns.total,
        helper_columns.checkbox_card,
        helper_columns.checkbox_cash,
        helper_columns.bank_income,
    }


def _clear_unselected_helper_values(
    sheet: Worksheet,
    row_number: int,
    helper_columns: HelperColumnMapping,
) -> None:
    selected_columns = _selected_helper_columns(helper_columns)

    for column in range(
        MIN_HELPER_COLUMN,
        MAX_HELPER_COLUMN + 1,
    ):
        if column in selected_columns:
            continue

        sheet.cell(
            row=row_number,
            column=column,
        ).value = None


def _apply_helper_total_grid(
    sheet: Worksheet,
    row_number: int,
    helper_columns: HelperColumnMapping,
) -> None:
    thin_side = Side(
        style="thin",
        color="FF000000",
    )
    full_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for column in _selected_helper_columns(helper_columns):
        sheet.cell(
            row=row_number,
            column=column,
        ).border = copy(full_border)


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return None


def _index_rows_by_date(sheet: Worksheet) -> dict[date, int]:
    rows_by_date: dict[date, int] = {}

    for row_number in range(1, sheet.max_row + 1):
        transaction_date = _as_date(
            sheet.cell(
                row=row_number,
                column=DATE_COLUMN,
            ).value
        )

        if transaction_date is not None:
            rows_by_date[transaction_date] = row_number

    return rows_by_date


def _validate_single_month(
    entries: list[DailyIncomeBookEntry],
) -> tuple[int, int] | None:
    periods = {(entry.date.year, entry.date.month) for entry in entries}

    if len(periods) > 1:
        raise IncomeBookExportError("entries must belong to one calendar month")
    return next(iter(periods), None)


def _clear_reserved_blank_cell(
    sheet: Worksheet,
    row_number: int,
) -> None:
    cell = sheet.cell(
        row=row_number,
        column=RESERVED_BLANK_COLUMN,
    )
    cell.value = None
    cell.style = "Normal"


def _write_daily_entry(
    sheet: Worksheet,
    row_number: int,
    entry: DailyIncomeBookEntry,
    helper_columns: HelperColumnMapping,
) -> None:
    zero = Decimal("0.00")

    sheet.cell(
        row=row_number,
        column=DATE_COLUMN,
    ).value = entry.date

    sheet.cell(
        row=row_number,
        column=INCOME_COLUMN,
    ).value = entry.total_income

    sheet.cell(
        row=row_number,
        column=REFUND_COLUMN,
    ).value = zero

    sheet.cell(
        row=row_number,
        column=ADJUSTED_INCOME_COLUMN,
    ).value = f"=B{row_number}-C{row_number}"

    sheet.cell(
        row=row_number,
        column=FREE_GOODS_COLUMN,
    ).value = zero

    sheet.cell(
        row=row_number,
        column=TOTAL_INCOME_COLUMN,
    ).value = f"=D{row_number}+E{row_number}"

    sheet.cell(
        row=row_number,
        column=SPECIAL_INCOME_TYPE_COLUMN,
    ).value = zero

    sheet.cell(
        row=row_number,
        column=SPECIAL_INCOME_AMOUNT_COLUMN,
    ).value = zero

    _clear_reserved_blank_cell(
        sheet,
        row_number,
    )

    _clear_unselected_helper_values(
        sheet,
        row_number,
        helper_columns,
    )

    card_letter = get_column_letter(helper_columns.checkbox_card)
    cash_letter = get_column_letter(helper_columns.checkbox_cash)
    bank_letter = get_column_letter(helper_columns.bank_income)

    sheet.cell(
        row_number,
        column=helper_columns.total,
    ).value = (
        f"={card_letter}{row_number}"
        f"+{cash_letter}{row_number}"
        f"+{bank_letter}{row_number}"
    )

    sheet.cell(
        row=row_number,
        column=helper_columns.checkbox_card,
    ).value = entry.checkbox_card_income

    sheet.cell(
        row=row_number,
        column=helper_columns.checkbox_cash,
    ).value = entry.checkbox_cash_income

    sheet.cell(
        row=row_number,
        column=helper_columns.bank_income,
    ).value = entry.bank_income


def _find_label_row(
    sheet: Worksheet,
    label: str,
) -> int:
    expected = label.strip().casefold()

    for row_number in range(1, sheet.max_row + 1):
        value = sheet.cell(
            row=row_number,
            column=DATE_COLUMN,
        ).value

        if isinstance(value, str) and value.strip().casefold() == expected:
            return row_number

    raise IncomeBookExportError(f"income-book row not found: {label}")


def _month_from_total_label(value: str) -> int | None:
    normalized = value.replace("\xa0", " ")
    normalized = normalized.strip().casefold()
    normalized = normalized.removesuffix(":")
    words = normalized.split()

    if not words:
        return None

    for month, month_name in MONTH_NAMES_UKR.items():
        if words[-1] == month_name:
            return month

    return None


def _index_month_total_rows(
    sheet: Worksheet,
) -> dict[int, int]:
    result: dict[int, int] = {}

    for row_number in range(1, sheet.max_row + 1):
        label_cell = sheet.cell(
            row=row_number,
            column=DATE_COLUMN,
        )
        value = label_cell.value

        if not isinstance(value, str):
            continue

        month = _month_from_total_label(value)

        if month is None:
            continue

        income_total = sheet.cell(
            row=row_number,
            column=INCOME_COLUMN,
        ).value

        if income_total is None:
            continue

        if month in result:
            month_name = MONTH_NAMES_UKR[month]
            raise IncomeBookExportError(
                f"multiple total rows found for month: {month_name}"
            )

        result[month] = row_number
        label_cell.value = f"Всього {MONTH_NAMES_UKR[month]}:"

    return result


def _validate_existing_month_total_rows(
    rows_by_date: dict[date, int],
    month_total_rows: dict[int, int],
    year: int,
) -> None:
    months_with_data = {
        transaction_date.month
        for transaction_date in rows_by_date
        if transaction_date.year == year
    }

    missing_months = sorted(months_with_data - month_total_rows.keys())

    if not missing_months:
        return

    missing_names = ", ".join(MONTH_NAMES_UKR[month] for month in missing_months)

    raise IncomeBookExportError(
        f"month total row not found for existing data: {missing_names}"
    )


def _period_month_total_rows(
    month_total_rows: dict[int, int],
    months_with_data: set[int],
    months: range,
) -> list[int]:
    period_months = [month for month in months if month in months_with_data]

    missing_months = [month for month in period_months if month not in month_total_rows]

    if missing_months:
        missing_names = ", ".join(MONTH_NAMES_UKR[month] for month in missing_months)

        raise IncomeBookExportError(
            f"month total row not found for existing data: {missing_names}"
        )

    return [month_total_rows[month] for month in period_months]


def _copy_row_style(
    sheet: Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height

    for column in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(
            row=source_row,
            column=column,
        )

        target_cell = sheet.cell(
            row=target_row,
            column=column,
        )

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)


def _write_derived_total_columns(
    sheet: Worksheet,
    row_number: int,
) -> None:
    sheet.cell(
        row=row_number,
        column=ADJUSTED_INCOME_COLUMN,
    ).value = f"=B{row_number}-C{row_number}"

    sheet.cell(
        row=row_number,
        column=TOTAL_INCOME_COLUMN,
    ).value = f"=D{row_number}+E{row_number}"

    sheet.cell(
        row=row_number,
        column=SPECIAL_INCOME_TYPE_COLUMN,
    ).value = Decimal("0.00")

    _clear_reserved_blank_cell(
        sheet,
        row_number,
    )


def _write_month_total(
    sheet: Worksheet,
    row_number: int,
    month: int,
    first_daily_row: int,
    last_daily_row: int,
    helper_columns: HelperColumnMapping,
) -> None:
    sheet.cell(
        row=row_number,
        column=DATE_COLUMN,
    ).value = f"Всього {MONTH_NAMES_UKR[month]}:"

    _clear_unselected_helper_values(
        sheet,
        row_number,
        helper_columns,
    )

    for column in _totalled_columns(helper_columns):
        letter = get_column_letter(column)
        sheet.cell(
            row=row_number,
            column=column,
        ).value = f"=SUM({letter}{first_daily_row}:{letter}{last_daily_row})"

    _write_derived_total_columns(sheet, row_number)

    _apply_helper_total_grid(
        sheet,
        row_number,
        helper_columns,
    )


def _write_period_total(
    sheet: Worksheet,
    row_number: int,
    label: str,
    month_total_rows: list[int],
    helper_columns: HelperColumnMapping,
) -> None:
    sheet.cell(
        row=row_number,
        column=DATE_COLUMN,
    ).value = label

    _clear_unselected_helper_values(
        sheet,
        row_number,
        helper_columns,
    )

    for column in _totalled_columns(helper_columns):
        letter = get_column_letter(column)
        references = "+".join(f"{letter}{month_row}" for month_row in month_total_rows)

        sheet.cell(
            row=row_number,
            column=column,
        ).value = f"={references}"

    _write_derived_total_columns(sheet, row_number)

    _apply_helper_total_grid(
        sheet,
        row_number,
        helper_columns,
    )


def _append_new_month(
    sheet: Worksheet,
    entries: list[DailyIncomeBookEntry],
    period: tuple[int, int],
    rows_by_date: dict[date, int],
    helper_columns: HelperColumnMapping,
) -> None:
    year, month = period

    year_total_row = _find_label_row(
        sheet,
        f"Всього {year} рік:",
    )

    latest_existing_date = max(rows_by_date)
    data_style_row = rows_by_date[latest_existing_date]

    month_total_rows = _index_month_total_rows(sheet)

    _validate_existing_month_total_rows(
        rows_by_date,
        month_total_rows,
        year,
    )

    months_with_data = {
        transaction_date.month
        for transaction_date in rows_by_date
        if transaction_date.year == year
    }

    if not month_total_rows:
        raise IncomeBookExportError("previous month total row not found")

    total_style_row = max(month_total_rows.values())

    summary_count = 1
    summary_count += int(month % 3 == 0)
    summary_count += int(month == 6)

    inserted_rows = len(entries) + summary_count

    sheet.insert_rows(
        year_total_row,
        amount=inserted_rows,
    )

    first_daily_row = year_total_row

    for offset, entry in enumerate(entries):
        row_number = first_daily_row + offset
        _copy_row_style(sheet, data_style_row, row_number)
        _write_daily_entry(sheet, row_number, entry, helper_columns)

    last_daily_row = first_daily_row + len(entries) - 1
    month_total_row = last_daily_row + 1

    _copy_row_style(
        sheet,
        total_style_row,
        month_total_row,
    )
    _write_month_total(
        sheet,
        month_total_row,
        month,
        first_daily_row,
        last_daily_row,
        helper_columns,
    )

    month_total_rows[month] = month_total_row
    months_with_data.add(month)
    next_row = month_total_row + 1

    if month % 3 == 0:
        quarter = (month - 1) // 3 + 1
        quarter_start = month - 2

        quarter_month_rows = _period_month_total_rows(
            month_total_rows,
            months_with_data,
            range(quarter_start, month + 1),
        )

        _copy_row_style(sheet, total_style_row, next_row)
        _write_period_total(
            sheet,
            next_row,
            f"Всього {quarter} кв {year}:",
            quarter_month_rows,
            helper_columns,
        )
        next_row += 1

    if month == 6:
        half_year_rows = _period_month_total_rows(
            month_total_rows,
            months_with_data,
            range(1, 7),
        )

        _copy_row_style(sheet, total_style_row, next_row)
        _write_period_total(
            sheet,
            next_row,
            f"Всього 1 півріччя {year}:",
            half_year_rows,
            helper_columns,
        )

    shifted_year_total_row = year_total_row + inserted_rows

    year_month_rows = _period_month_total_rows(
        month_total_rows,
        months_with_data,
        range(1, month + 1),
    )

    _write_period_total(
        sheet,
        shifted_year_total_row,
        f"Всього {year} рік:",
        year_month_rows,
        helper_columns,
    )


def export_income_book(
    template_path: Path,
    output_path: Path,
    entries: list[DailyIncomeBookEntry],
    *,
    sheet_name: str,
    helper_columns: HelperColumnMapping | None = None,
) -> Path:
    if helper_columns is None:
        helper_columns = HelperColumnMapping()

    if template_path.resolve() == output_path.resolve():
        raise IncomeBookExportError("output path must differ from template path")

    period = _validate_single_month(entries)
    entries = sorted(entries, key=lambda entry: entry.date)

    try:
        workbook = load_workbook(
            template_path,
            data_only=False,
        )
    except Exception as error:
        raise IncomeBookExportError(
            f"can't read income-book template '{template_path.name}'"
        ) from error

    try:
        if sheet_name not in workbook.sheetnames:
            raise MissingIncomeBookSheetError(
                f"income-book sheet not found: {sheet_name}"
            )

        sheet = workbook[sheet_name]
        rows_by_date = _index_rows_by_date(sheet)

        missing_dates = sorted(
            entry.date for entry in entries if entry.date not in rows_by_date
        )

        if missing_dates:
            latest_existing_date = max(rows_by_date, default=None)

            latest_period = (
                (
                    latest_existing_date.year,
                    latest_existing_date.month,
                )
                if latest_existing_date is not None
                else None
            )

            if (
                period is not None
                and latest_period is not None
                and period > latest_period
            ):
                _append_new_month(
                    sheet,
                    entries,
                    period,
                    rows_by_date,
                    helper_columns,
                )

            else:
                formatted_dates = ", ".join(
                    missing_date.isoformat() for missing_date in missing_dates
                )
                raise MissingIncomeBookDateError(
                    f"income-book dates not found: {formatted_dates}"
                )
        else:
            for entry in entries:
                row_number = rows_by_date[entry.date]
                _write_daily_entry(sheet, row_number, entry, helper_columns)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )
        workbook.save(output_path)

    finally:
        workbook.close()

    return output_path
