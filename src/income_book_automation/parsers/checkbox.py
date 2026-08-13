"""Parser for Checkbox Z-report workbooks."""

from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from income_book_automation.models import DailyCheckboxRevenue

REQUIRED_HEADERS = {
    "opened_at": "Дата відкриття",
    "card_revenue": "Виручка безготівка",
    "card_refund": "Повернення безготівка",
    "cash_revenue": "Виручка готівка",
    "cash_refund": "Повернення готівка",
}


class CheckboxParseError(Exception):
    """Raised when a Checkbox workbook has any error."""


class CheckboxFormatError(CheckboxParseError):
    """Raised when a Checkbox workbook has an unexpected structure."""


class MissingCheckboxColumnError(CheckboxFormatError):
    """Raised when a required Checkbox column is missing."""

    def __init__(
        self,
        filename: str,
        missing_headers: tuple[str, ...],
    ) -> None:
        self.filename = filename
        self.missing_headers = missing_headers

        columns = ", ".join(f"«{header}»" for header in missing_headers)

        if len(missing_headers) == 1:
            message = (
                f"У Z-звіті Checkbox «{filename}» відсутня обов’язкова "
                f"колонка: {columns}. Книгу доходів не сформовано."
            )
        else:
            message = (
                f"У Z-звіті Checkbox «{filename}» відсутні обов’язкові "
                f"колонки: {columns}. Книгу доходів не сформовано."
            )

        super().__init__(message)


class InvalidCheckboxRowError(CheckboxParseError):
    """Raised when a checkbox workbook has a different value."""


def normalize_header(value: object) -> str:
    if value is None:
        return ""

    text = str(value).replace("\u00a0", " ")
    return " ".join(text.split()).casefold()


def resolve_column_headers(
    header_row: tuple[object, ...], path: Path
) -> dict[str, int]:
    available_headers = {
        normalize_header(header): index
        for index, header in enumerate(header_row)
        if header is not None
    }

    indexes: dict[str, int] = {}
    missing_headers: list[str] = []

    for field_name, required_header in REQUIRED_HEADERS.items():
        normalized_header = normalize_header(required_header)

        if normalized_header not in available_headers:
            missing_headers.append(required_header)
            continue
        indexes[field_name] = available_headers[normalized_header]

    if missing_headers:
        raise MissingCheckboxColumnError(
            path.name,
            tuple(missing_headers),
        )

    return indexes


def is_blank_value(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def parse_checkbox_row(
    row: tuple[object, ...],
    column_indexes: dict[str, int],
    path: Path,
    row_number: int,
) -> DailyCheckboxRevenue:
    opened_at = row[column_indexes["opened_at"]]
    if is_blank_value(opened_at):
        raise InvalidCheckboxRowError(
            f"File '{path.name}', row {row_number}, "
            f"column '{REQUIRED_HEADERS['opened_at']}': "
            "required value is missing"
        )
    try:
        opened_date = opened_at.date()
    except AttributeError as error:
        raise InvalidCheckboxRowError(
            f"File '{path.name}', row {row_number}, "
            f"column '{REQUIRED_HEADERS['opened_at']}': "
            f"invalid value {opened_at!r}"
        ) from error

    card_revenue_value = row[column_indexes["card_revenue"]]
    card_revenue = parse_decimal_value(
        card_revenue_value,
        path,
        row_number,
        REQUIRED_HEADERS["card_revenue"],
    )

    card_refund_value = row[column_indexes["card_refund"]]
    card_refund = parse_decimal_value(
        card_refund_value,
        path,
        row_number,
        REQUIRED_HEADERS["card_refund"],
    )

    cash_revenue_value = row[column_indexes["cash_revenue"]]
    cash_revenue = parse_decimal_value(
        cash_revenue_value,
        path,
        row_number,
        REQUIRED_HEADERS["cash_revenue"],
    )

    cash_refund_value = row[column_indexes["cash_refund"]]
    cash_refund = parse_decimal_value(
        cash_refund_value,
        path,
        row_number,
        REQUIRED_HEADERS["cash_refund"],
    )

    try:
        return DailyCheckboxRevenue(
            date=opened_date,
            card_revenue=card_revenue,
            card_refund=card_refund,
            cash_revenue=cash_revenue,
            cash_refund=cash_refund,
        )
    except ValidationError as error:
        raise InvalidCheckboxRowError(
            f"File '{path.name}', row {row_number}: invalid monetary values"
        ) from error


def parse_decimal_value(
    value: object, path: Path, row_number: int, column_name: str
) -> Decimal:
    if is_blank_value(value):
        raise InvalidCheckboxRowError(
            f"File '{path.name}', row {row_number}, "
            f"column '{column_name}': required value is missing"
        )

    try:
        return Decimal(str(value).strip())
    except InvalidOperation as error:
        raise InvalidCheckboxRowError(
            f"File '{path.name}', row {row_number}, "
            f"column '{column_name}': "
            f"invalid value {value!r}"
        ) from error


def validate_formula_results(
    value_row: tuple[object, ...],
    formula_row: tuple[object, ...],
    column_indexes: dict[str, int],
    path: Path,
    row_number: int,
) -> None:
    for field_name, column_index in column_indexes.items():
        formula_cell = formula_row[column_index]
        calculated_value = value_row[column_index]

        if getattr(formula_cell, "data_type", None) == "f" and is_blank_value(
            calculated_value
        ):
            raise InvalidCheckboxRowError(
                f"File '{path.name}', row {row_number}, "
                f"column '{REQUIRED_HEADERS[field_name]}': "
                "formula has no cached result"
            )


def parse_checkbox_file(path: Path) -> list[DailyCheckboxRevenue]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            formula_workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
            )
        except Exception as error:
            workbook.close()

            raise CheckboxParseError(
                f"Can't read Checkbox formulas from file: '{path.name}'."
            ) from error
    except Exception as error:
        raise CheckboxParseError(
            f"Can't read Checkbox file: '{path.name}' the file is corrupted or has an invalid format."
        ) from error

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)

        formula_worksheet = formula_workbook.active
        formula_rows = formula_worksheet.iter_rows(values_only=False)

        header_row = next(rows, None)
        next(formula_rows, None)

        if header_row is None:
            raise CheckboxFormatError("Checkbox workbook is empty.")

        column_indexes = resolve_column_headers(header_row, path)

        records: list[DailyCheckboxRevenue] = []

        required_indexes = tuple(column_indexes.values())

        for row_number, (row, formula_row) in enumerate(
            zip(rows, formula_rows, strict=True),
            start=2,
        ):
            validate_formula_results(
                row,
                formula_row,
                column_indexes,
                path,
                row_number,
            )

            required_values = tuple(row[index] for index in required_indexes)

            if all(is_blank_value(value) for value in required_values):
                continue

            record = parse_checkbox_row(
                row,
                column_indexes,
                path,
                row_number,
            )
            records.append(record)

        return records

    finally:
        workbook.close()
        formula_workbook.close()
